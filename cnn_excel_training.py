import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

# ---------------------------
# Helper functions for our CNN:
# ---------------------------

def conv2d(x, kernel):
    # x is 3x3, kernel is 2x2; output is (3-2+1)x(3-2+1)=2x2
    H, W = x.shape
    kH, kW = kernel.shape
    outH, outW = H - kH + 1, W - kW + 1
    out = np.zeros((outH, outW))
    for i in range(outH):
        for j in range(outW):
            region = x[i:i+kH, j:j+kW]
            out[i, j] = np.sum(region * kernel)
    return out

def relu(x):
    return np.maximum(0, x)

def flatten(x):
    return x.flatten()

def dense_forward(a, W, b):
    # a: shape (4,), W: shape (4,7), b: shape (7,)
    z = np.dot(a, W) + b
    return z

def softmax(z):
    expz = np.exp(z - np.max(z))
    return expz / np.sum(expz)

def cross_entropy_loss(p, t):
    # p: predicted probability vector, t: one-hot target vector
    # Loss = -sum(t*log(p)). t is one-hot so equals -log(probability for correct class)
    return -np.log(p[np.argmax(t)] + 1e-8)

# ---------------------------
# Define our fixed filter and initial Dense parameters:
# ---------------------------
conv_filter = np.array([[1, 1],
                        [0,  1]], dtype=np.float32)

# We'll use the same initial Dense parameters as in the workbook:
W_init = np.array([
    [0.5,  0.2, -0.3,  0.7,  0.1, -0.4,  0.3],
    [0.1, -0.1,  0.4,  0.2, -0.5,  0.3,  0.6],
    [-0.2, 0.3,  0.5, -0.1,  0.4,  0.2, -0.3],
    [0.7, -0.6,  0.2,  0.1,  0.3,  0.5,  0.2]
], dtype=np.float32)   # shape (4,7)
b_init = np.array([0.1, -0.1, 0.0, 0.2, -0.2, 0.1, 0.0], dtype=np.float32)  # shape (7,)

# ---------------------------
# Define our training data: the 7 Tetris pieces
# ---------------------------
# Each piece is a 3x3 matrix:
piece1 = np.array([[0,1,0],
                   [0,1,0],
                   [0,1,0]], dtype=np.float32)  # Slinky Snake (I)
piece2 = np.array([[1,1,0],
                   [1,1,0],
                   [0,0,0]], dtype=np.float32)  # Bubble Block (O)
piece3 = np.array([[0,1,0],
                   [1,1,1],
                   [0,0,0]], dtype=np.float32)  # Topsy Turvy (T)
piece4 = np.array([[0,1,1],
                   [1,1,0],
                   [0,0,0]], dtype=np.float32)  # Slippery Slide (S)
piece5 = np.array([[1,1,0],
                   [0,1,1],
                   [0,0,0]], dtype=np.float32)  # Zany Zigzag (Z)
piece6 = np.array([[1,0,0],
                   [1,1,1],
                   [0,0,0]], dtype=np.float32)  # Jolly Jumper (J)
piece7 = np.array([[0,0,1],
                   [1,1,1],
                   [0,0,0]], dtype=np.float32)  # Lively L (L)

pieces = [piece1, piece2, piece3, piece4, piece5, piece6, piece7]
# Targets: one-hot encoding (7 classes)
targets = np.eye(7, dtype=np.float32)  # row i is the target for piece i+1

# ---------------------------
# Training parameters:
# ---------------------------
learning_rate = 0.1
num_epochs = 5000

# Initialize trainable parameters with initial values
W = W_init.copy()
b = b_init.copy()
# We'll train by iterating over all 7 examples and averaging gradients.
for epoch in range(num_epochs):
    grad_W = np.zeros_like(W)
    grad_b = np.zeros_like(b)
    total_loss = 0.0
    for x, t in zip(pieces, targets):
        # Forward pass:
        conv_out = conv2d(x, conv_filter)       # shape 2x2
        relu_out = relu(conv_out)                # shape 2x2
        a = flatten(relu_out)                    # shape (4,)
        z = dense_forward(a, W, b)               # shape (7,)
        p = softmax(z)                         # shape (7,)
        loss = cross_entropy_loss(p, t)
        total_loss += loss
        # Backprop: gradient w.r.t. dense output: (p - t)
        dz = p - t                           # shape (7,)
        # Gradients:
        # dW = a[:, None] * dz[None, :]
        grad_W += np.outer(a, dz)              # shape (4,7)
        grad_b += dz                         # shape (7,)
    # Average gradients over 7 examples
    grad_W /= 7.0
    grad_b /= 7.0
    # Update parameters:
    W -= learning_rate * grad_W
    b -= learning_rate * grad_b
    if (epoch+1) % 500 == 0:
        avg_loss = total_loss / 7.0
        print(f"Epoch {epoch+1}, Loss: {avg_loss:.4f}")

print("Training complete.")
print("Optimized Dense Weights:")
print(W)
print("Optimized Dense Biases:")
print(b)

# ------------------------------------
# Now create the Excel sheet:
# ------------------------------------
# Create a new workbook and remove the default sheet
wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

bold_font = Font(bold=True)

##########################################
# 1. Inputs sheet (Pre-defined pieces)  #
##########################################
ws_inputs = wb.create_sheet(title="Inputs")
# Add bold column headers in row 1 (columns B-D)
ws_inputs.cell(row=1, column=2, value="Col1").font = bold_font
ws_inputs.cell(row=1, column=3, value="Col2").font = bold_font
ws_inputs.cell(row=1, column=4, value="Col3").font = bold_font

# Define pieces (3x3 matrices), cute names, and colors.
# These are the same as before:
cute_names = [
    "Slinky Snake", "Bubble Block", "Topsy Turvy",
    "Slippery Slide", "Zany Zigzag", "Jolly Jumper", "Lively L"
]
tetris_colors = [
    "00FFFF",  # Cyan
    "FFFF00",  # Yellow
    "800080",  # Purple
    "00FF00",  # Green
    "FF0000",  # Red
    "0000FF",  # Blue
    "FFA500"   # Orange
]

start_row = 2  # start at row 2 so header remains at row 1
for idx, piece in enumerate(pieces, start=1):
    fill = PatternFill(start_color=tetris_colors[idx-1],
                       end_color=tetris_colors[idx-1],
                       fill_type="solid")
    for i, row_vals in enumerate(piece, start=1):
        for j, val in enumerate(row_vals, start=1):
            cell = ws_inputs.cell(row=start_row + i, column=j+1, value=val)
            if val == 1:
                cell.fill = fill
        ws_inputs.cell(row=start_row + i, column=1, value=f"Row{i}").font = bold_font
    # Write the cute name in column F and instruction in column F
    ws_inputs.cell(row=start_row+1, column=6, value=cute_names[idx-1]).font = bold_font
    ws_inputs.cell(row=start_row+2, column=6, value="(Manually copy this piece to Inference)")
    start_row += 5

#################################
# 2. Filter sheet (Conv kernel) #
#################################
ws_filter = wb.create_sheet(title="Filter")
ws_filter.cell(row=1, column=1, value="FRow1").font = bold_font
ws_filter.cell(row=1, column=2, value="FRow2").font = bold_font
# Set filter to only use 0 and 1 values. For example, use a simple binary filter:
filter_vals = [
    [1, 1],
    [0, 1]
]
for i, row_vals in enumerate(filter_vals, start=1):
    for j, val in enumerate(row_vals, start=1):
        ws_filter.cell(row=i, column=j, value=val)

#################################################
# 3. Inference sheet (User input piece)         #
#################################################
ws_infer = wb.create_sheet(title="Inference")
ws_infer.cell(row=1, column=2, value="Col1").font = bold_font
ws_infer.cell(row=1, column=3, value="Col2").font = bold_font
ws_infer.cell(row=1, column=4, value="Col3").font = bold_font
ws_infer.cell(row=2, column=1, value="Row1").font = bold_font
ws_infer.cell(row=3, column=1, value="Row2").font = bold_font
ws_infer.cell(row=4, column=1, value="Row3").font = bold_font
# Preload with piece1 as default (user can modify these cells)
ws_infer["B2"] = 0
ws_infer["C2"] = 1
ws_infer["D2"] = 0
ws_infer["B3"] = 0
ws_infer["C3"] = 1
ws_infer["D3"] = 0
ws_infer["B4"] = 0
ws_infer["C4"] = 1
ws_infer["D4"] = 0

#################################################
# 4. Conv sheet: Compute convolution from Inference!B2:D4 #
#################################################
ws_conv = wb.create_sheet(title="Conv")
ws_conv.cell(row=1, column=1, value="Conv Output").font = bold_font
ws_conv["A2"] = "=SUMPRODUCT(Inference!B2:C3, Filter!A1:B2)"
ws_conv["B2"] = "=SUMPRODUCT(Inference!C2:D3, Filter!A1:B2)"
ws_conv["A3"] = "=SUMPRODUCT(Inference!B3:C4, Filter!A1:B2)"
ws_conv["B3"] = "=SUMPRODUCT(Inference!C3:D4, Filter!A1:B2)"

#################################################
# 5. ReLU sheet: Apply MAX(0, value)
#################################################
ws_relu = wb.create_sheet(title="ReLU")
ws_relu.cell(row=1, column=1, value="ReLU Output").font = bold_font
ws_relu["A2"] = "=MAX(0, Conv!A2)"
ws_relu["B2"] = "=MAX(0, Conv!B2)"
ws_relu["A3"] = "=MAX(0, Conv!A3)"
ws_relu["B3"] = "=MAX(0, Conv!B3)"

#################################################
# 6. Flatten sheet: Flatten the 2x2 ReLU result into a 4-vector
#################################################
ws_flat = wb.create_sheet(title="Flatten")
ws_flat.cell(row=1, column=1, value="Flattened Vector").font = bold_font
ws_flat["A2"] = "=ReLU!A2"
ws_flat["A3"] = "=ReLU!B2"
ws_flat["A4"] = "=ReLU!A3"
ws_flat["A5"] = "=ReLU!B3"

###################################################
# 7. DenseParameters sheet (Trainable parameters)
###################################################
ws_dense_params = wb.create_sheet(title="DenseParameters")
ws_dense_params.cell(row=1, column=1, value="Weights").font = bold_font
# Write the *trained* W and b values to the sheet.
for i, row_vals in enumerate(W, start=1):
    for j, val in enumerate(row_vals, start=1):
        ws_dense_params.cell(row=i, column=j, value=val)
ws_dense_params.cell(row=6, column=1, value="Biases").font = bold_font
for j, val in enumerate(b, start=1):
    ws_dense_params.cell(row=6, column=j, value=val)

#####################################
# 8. Dense sheet: Compute dense layer outputs (1x7 vector)
#####################################
ws_dense = wb.create_sheet(title="Dense")
ws_dense.cell(row=1, column=1, value="Dense Output").font = bold_font
for j in range(1, 8):
    col_letter = get_column_letter(j)
    formula = f"=SUMPRODUCT(Flatten!A2:A5, DenseParameters!{col_letter}1:{col_letter}4) + DenseParameters!{col_letter}6"
    ws_dense.cell(row=2, column=j, value=formula)

#####################################
# 9. Softmax sheet: Compute softmax probabilities (1x7 vector)
#####################################
ws_softmax = wb.create_sheet(title="Softmax")
ws_softmax.cell(row=1, column=1, value="Softmax Probabilities").font = bold_font
# Write label "Denom" in cell H1
ws_softmax.cell(row=1, column=8, value="Denom").font = bold_font
# Place denominator formula in cell H2
denom_formula = "=EXP(Dense!A2)+EXP(Dense!B2)+EXP(Dense!C2)+EXP(Dense!D2)+EXP(Dense!E2)+EXP(Dense!F2)+EXP(Dense!G2)"
ws_softmax["H2"] = denom_formula
for j in range(1, 8):
    col_letter = get_column_letter(j)
    softmax_formula = f"=EXP(Dense!{col_letter}2)/($H$2)"
    ws_softmax.cell(row=2, column=j, value=softmax_formula)

#####################################
# 10. Classification sheet: Display predicted cute name.
#####################################
ws_class = wb.create_sheet(title="Classification")
ws_class.cell(row=1, column=1, value="Predicted Piece:").font = bold_font
ws_class.cell(row=1, column=2,
    value='=INDEX({"Slinky Snake","Bubble Block","Topsy Turvy","Slippery Slide","Zany Zigzag","Jolly Jumper","Lively L"},MATCH(MAX(Softmax!A2:G2),Softmax!A2:G2,0))'
).font = bold_font

#####################################
# Save the workbook
excel_filename = "cnn_inference_trained.xlsx"
wb.save(excel_filename)
print(f"Excel file '{excel_filename}' created with trained DenseParameters for inference.")